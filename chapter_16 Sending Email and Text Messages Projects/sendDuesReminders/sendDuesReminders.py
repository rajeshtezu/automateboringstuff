#! python3
# sendDuesReminders.py - Sends emails based on payment status in spreadsheet.

import openpyxl, smtplib, sys, getpass

# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.get_highest_column()
latestMonth = sheet.cell(row=1, column=lastCol).value

# Check each member's payment status.
unpaidMembers = {}
for r in range(2, sheet.get_highest_row() + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email
        
# Log in to email account.
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()

emailAddr = input('Your Email: ')
while not validateEmail(emailAddr):
    print('Invalid email')
    emailAddr = input('Enter valid Email: ')
    
password = getpass.getpass('Password: ')

smtpObj.login(emailAddr, password)


# Send out reminder emails.
for name, email in unpaidMembers.items():
    body = '''Subject: %s dues unpaid.\nDear %s, \nRecords show that you have not paid dues for %s. Please 
              make this payment as soon as possible. Thank you!''' % (latestMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    
    sendmailStatus = smtpObj.sendmail(emailAddr, email, body)
    
    if sendmailStatus != {}:
        print('Threre was a problem sending email to %s: %s' % (email, sendmailStatus))
        
smtpObj.quit()        
    

def validateEmail(mail):
    emailRegex = re.compile(r'''(
     [a-zA-Z0-9._%+-]+      # username
     @                      # @ symbol
     [a-zA-Z0-9.-]+         # domain name
     (\.[a-zA-Z]{2,4})      # dot-something
     )''', re.VERBOSE)
    
    if emailRegex.search(mail):
        return True
    else:
        return False