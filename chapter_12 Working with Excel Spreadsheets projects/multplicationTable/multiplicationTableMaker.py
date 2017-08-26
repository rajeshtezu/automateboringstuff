
#! /usr/bin/python3

'''
A program that takes a number N from the command line and creates an N×N multiplication table in an Excel spreadsheet.
'''

import openpyxl
from openpyxl.styles import Font
import sys

if len(sys.argv) < 2:
	sys.exit('Usage: python3 multiplicationTableMaker.py <number>')

number = int(sys.argv[1])

def multiplicationTable(N):
	wb = openpyxl.Workbook()
	sheet = wb.get_sheet_by_name('Sheet')

	lableFont = Font(bold=True)

	print('Creating Multiplication Table...')
	for i in range(1, N+1):
		for j in range(1, N+1):
			# Filling column heading
			sheet.cell(row=1, column=i+1).value = i
			sheet.cell(row=1, column=i+1).font = lableFont

			# Filling row heading
			sheet.cell(row=i+1, column=1).value = i
			sheet.cell(row=i+1, column=1).font = lableFont

			# Filling rectangular box
			sheet.cell(row=i+1, column=j+1).value = i*j

	wb.save('multiplicationTable.xlsx')
	print('Done.')


if __name__ == '__main__':
	multiplicationTable(number)
